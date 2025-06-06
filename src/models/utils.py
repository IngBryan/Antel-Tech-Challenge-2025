import csv
import os
from openpyxl import load_workbook
import xlrd


def excel_a_csv(ruta_entrada, ruta_salida):
    extension = os.path.splitext(ruta_entrada)[1].lower()

    if extension == ".xlsx":
        wb = load_workbook(ruta_entrada, data_only=True)
        hoja = wb.active

        primera_fila = next(hoja.iter_rows(values_only=True))
        max_columnas = len([celda for celda in primera_fila if celda is not None])

        with open(ruta_salida, "w", newline="", encoding="utf-8") as archivo_csv:
            escritor = csv.writer(archivo_csv)
            for fila in hoja.iter_rows(values_only=True):
                if any(celda not in (None, '') for celda in fila[:max_columnas]):
                    escritor.writerow(fila[:max_columnas])

    elif extension == ".xls":
        wb = xlrd.open_workbook(ruta_entrada)
        hoja = wb.sheet_by_index(0)

        max_columnas = hoja.ncols

        with open(ruta_salida, "w", newline="", encoding="utf-8") as archivo_csv:
            escritor = csv.writer(archivo_csv)
            for fila_idx in range(hoja.nrows):
                fila = hoja.row_values(fila_idx)
                if any(celda not in (None, '') for celda in fila[:max_columnas]):
                    escritor.writerow(fila[:max_columnas])

    else:
        raise ValueError("Extensión de archivo no soportada: debe ser .xls o .xlsx")